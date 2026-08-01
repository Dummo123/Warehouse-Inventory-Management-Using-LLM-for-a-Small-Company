"""
Импорт начальных данных из warehouse_report.xlsx.

Запуск:
    python scripts/seed_data.py --excel path/to/warehouse_report.xlsx

────────────────────────────────────────────────────────────────────────────
ИСПРАВЛЕНИЕ (30.07.2026): раньше весь скрипт работал в ОДНОЙ транзакции —
db.commit() вызывался только один раз, в самом конце. Из-за этого, если
что-то падало на последнем шаге (например, лист "Accounts Balances"
отсутствовал/был пустой в чьей-то копии файла, или БД была в неожиданном
состоянии), срабатывал except → db.rollback() — и он откатывал ВООБЩЕ ВСЁ,
включая уже "успешно" пропечатанные склады, артикулы, остатки и 500+
движений. Именно поэтому лог показывал сплошные "✅", а в самой БД потом
было 0 строк.

Теперь каждый смысловой блок (склады+артикулы+BOM, остатки, движения,
пользователи, admin) коммитится ОТДЕЛЬНО. Если что-то упадёт на шаге
"пользователи" — склады/артикулы/остатки/движения уже будут сохранены
и никуда не денутся.
────────────────────────────────────────────────────────────────────────────
"""
import argparse
import os
import re
import sys
import traceback
from datetime import datetime

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.core.config import settings
from app.core.security import get_password_hash
from app.db.session import SessionLocal, init_db
from app.models.models import (
    Article, ArticleType, BOM, Movement, MovementType,
    Stock, User, UserRole, Warehouse, WarehouseType,
)


# ── Вспомогательные функции (без изменений) ───────────────────────────────

def get_or_create_warehouse(db: Session, name: str, wh_type: WarehouseType) -> Warehouse:
    wh = db.query(Warehouse).filter(Warehouse.warehouse_type == wh_type).first()
    if not wh:
        wh = Warehouse(name=name, warehouse_type=wh_type)
        db.add(wh)
        db.flush()
    return wh


def get_or_create_article(db, code, name, art_type, cost=0.0, responsible=None, comment=None):
    a = db.query(Article).filter(Article.code == str(code)).first()
    if not a:
        a = Article(code=str(code), name=name, article_type=art_type,
                    cost_price=cost, responsible=responsible, comment=comment)
        db.add(a)
        db.flush()
    return a


def upsert_stock(db: Session, article_id: int, warehouse_id: int, quantity: float):
    s = db.query(Stock).filter_by(article_id=article_id, warehouse_id=warehouse_id).first()
    if s:
        s.quantity = quantity
    else:
        db.add(Stock(article_id=article_id, warehouse_id=warehouse_id, quantity=quantity))
    db.flush()


def parse_bom_string(bom_str: str) -> list[tuple[str, float]]:
    """'1132 (1), 1118 (1), 1124 (2)' → [('1132', 1.0), ...]"""
    result = []
    for match in re.finditer(r'(\w+)\s*\((\d+(?:\.\d+)?)\)', str(bom_str)):
        result.append((match.group(1), float(match.group(2))))
    return result


# ── Диагностика устаревшей схемы БД ────────────────────────────────────────

def check_legacy_schema(db: Session):
    """
    Предупреждает, если в БД остались таблицы/колонки от старой версии
    models.py (например, production_batches — удалён из кода, но мог
    остаться в уже существующем Docker-томе, если тот не пересоздавался).
    create_all() НЕ удаляет и НЕ изменяет уже существующие таблицы, поэтому
    после смены схемы старый volume нужно сбрасывать вручную.
    """
    try:
        legacy_table = db.execute(text(
            "SELECT to_regclass('public.production_batches')"
        )).scalar()
        legacy_column = db.execute(text(
            "SELECT column_name FROM information_schema.columns "
            "WHERE table_name='movements' AND column_name='production_batch_id'"
        )).scalar()
        if legacy_table or legacy_column:
            print(
                "\n⚠️  ВНИМАНИЕ: в базе данных обнаружены следы СТАРОЙ схемы "
                "(таблица production_batches и/или колонка movements.production_batch_id).\n"
                "   Это значит, что Docker-том с данными был создан на более ранней "
                "версии кода и не пересоздавался.\n"
                "   create_all() не удаляет старые объекты сам — рекомендуется сделать:\n"
                "       docker-compose down -v\n"
                "       docker-compose up -d db\n"
                "   и затем повторно запустить этот скрипт с нуля.\n"
            )
    except Exception:
        # Диагностика необязательна — если не получилось проверить (например,
        # БД совсем пустая и таблиц ещё нет), просто продолжаем импорт.
        db.rollback()


# ── Основной импорт ─────────────────────────────────────────────────────

def seed(excel_path: str):
    print(f"📂 Читаем файл: {excel_path}")
    if not os.path.exists(excel_path):
        print(f"❌ Файл не найден: {excel_path}")
        print("   Проверьте путь — используйте полный путь или скопируйте файл в папку smarttherm.")
        sys.exit(1)

    wb = openpyxl.load_workbook(excel_path)
    db: Session = SessionLocal()
    check_legacy_schema(db)

    # ── Блок 1: склады + артикулы + BOM (критичный, всё в одной транзакции) ──
    try:
        print("\n🏭 Создаём склады...")
        wh_comp = get_or_create_warehouse(db, "Склад компонентов и полуфабрикатов", WarehouseType.COMPONENTS)
        wh_fin = get_or_create_warehouse(db, "Склад готовой продукции", WarehouseType.FINISHED_GOODS)
        db.flush()
        print(f"   ✅ {wh_comp.name}")
        print(f"   ✅ {wh_fin.name}")

        print("\n📦 Импортируем полуфабрикаты...")
        articles_map: dict[str, Article] = {}
        if "Semi-Finished Products" not in wb.sheetnames:
            print("   ⚠️  Лист 'Semi-Finished Products' не найден в файле — пропускаем.")
        else:
            for row in wb["Semi-Finished Products"].iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                code = str(row[0]).strip()
                a = get_or_create_article(
                    db, code, str(row[1]).strip(),
                    ArticleType.COMPONENT,
                    float(row[2] or 0), str(row[3] or ""),
                    str(row[4]) if row[4] else None,
                )
                articles_map[code] = a
        db.flush()
        print(f"   ✅ {len(articles_map)} полуфабрикатов")

        print("\n🔧 Импортируем готовые изделия и BOM...")
        finished_map: dict[str, Article] = {}
        if "Products" not in wb.sheetnames:
            print("   ⚠️  Лист 'Products' не найден в файле — пропускаем.")
        else:
            for row in wb["Products"].iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                art_code = str(row[0]).strip()
                a = get_or_create_article(db, art_code, str(row[1]).strip(), ArticleType.FINISHED)
                finished_map[art_code] = a
                db.flush()

                db.query(BOM).filter(BOM.parent_id == a.id).delete()
                seen = set()
                for comp_code, qty in parse_bom_string(str(row[2] or "")):
                    if comp_code in seen:
                        continue
                    seen.add(comp_code)
                    comp = articles_map.get(comp_code) or db.query(Article).filter(Article.code == comp_code).first()
                    if comp:
                        db.add(BOM(parent_id=a.id, child_id=comp.id, quantity=qty))
                    else:
                        print(f"   ⚠️  BOM: компонент '{comp_code}' не найден для {art_code}")
        db.flush()
        print(f"   ✅ {len(finished_map)} готовых изделий")

        db.commit()  # ← ЧЕКПОИНТ 1: склады, артикулы и BOM сохранены безвозвратно
        print("   💾 Сохранено (checkpoint 1/4)")

    except Exception:
        db.rollback()
        print("\n❌ Ошибка на шаге 'склады/артикулы/BOM':")
        traceback.print_exc()
        db.close()
        sys.exit(1)

    # ── Блок 2: остатки (критичный) ────────────────────────────────────────
    try:
        print("\n📊 Устанавливаем остатки...")
        count_stock = 0
        if "Stock" not in wb.sheetnames:
            print("   ⚠️  Лист 'Stock' не найден в файле — пропускаем.")
        else:
            for row in wb["Stock"].iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                code, name = str(row[0]).strip(), str(row[1]).strip()
                qty, cost = float(row[2] or 0), float(row[3] or 0)

                a = db.query(Article).filter(Article.code == code).first()
                if not a:
                    a = get_or_create_article(db, code, name, ArticleType.COMPONENT, cost)
                a.cost_price = cost / qty if qty > 0 else cost

                wh = wh_fin if a.article_type == ArticleType.FINISHED else wh_comp
                upsert_stock(db, a.id, wh.id, qty)
                count_stock += 1

        for a in list(articles_map.values()) + list(finished_map.values()):
            wh = wh_fin if a.article_type == ArticleType.FINISHED else wh_comp
            if not db.query(Stock).filter_by(article_id=a.id, warehouse_id=wh.id).first():
                db.add(Stock(article_id=a.id, warehouse_id=wh.id, quantity=0))
        db.flush()
        print(f"   ✅ {count_stock} записей остатков")

        db.commit()  # ← ЧЕКПОИНТ 2
        print("   💾 Сохранено (checkpoint 2/4)")

    except Exception:
        db.rollback()
        print("\n❌ Ошибка на шаге 'остатки' (склады/артикулы/BOM из checkpoint 1 УЖЕ сохранены и не пострадали):")
        traceback.print_exc()
        db.close()
        sys.exit(1)

    # ── Блок 3: движения (критичный) ──────────────────────────────────────
    try:
        print("\n📋 Импортируем историю движений...")
        count_mv, skipped = 0, 0
        if "Movements" not in wb.sheetnames:
            print("   ⚠️  Лист 'Movements' не найден в файле — пропускаем.")
        else:
            for row in wb["Movements"].iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                date_val = row[0]
                name = str(row[1]).strip()
                incoming = float(row[2] or 0)
                outgoing = float(row[3] or 0)
                comment = str(row[4]) if row[4] else None

                mv_date = date_val if isinstance(date_val, datetime) else datetime.utcnow()

                a = db.query(Article).filter(Article.name.ilike(f"%{name}%")).first()
                if not a:
                    skipped += 1
                    continue

                wh = wh_fin if a.article_type == ArticleType.FINISHED else wh_comp
                if incoming > 0:
                    db.add(Movement(article_id=a.id, warehouse_id=wh.id,
                                    movement_type=MovementType.RECEIPT,
                                    quantity=incoming, comment=comment, movement_date=mv_date))
                    count_mv += 1
                if outgoing > 0:
                    db.add(Movement(article_id=a.id, warehouse_id=wh.id,
                                    movement_type=MovementType.SHIPMENT,
                                    quantity=outgoing, comment=comment, movement_date=mv_date))
                    count_mv += 1
        db.flush()
        print(f"   ✅ {count_mv} движений импортировано, пропущено: {skipped}")

        db.commit()  # ← ЧЕКПОИНТ 3
        print("   💾 Сохранено (checkpoint 3/4)")

    except Exception:
        db.rollback()
        print("\n❌ Ошибка на шаге 'движения' (склады/артикулы/BOM/остатки из checkpoint 1-2 УЖЕ сохранены):")
        traceback.print_exc()
        db.close()
        sys.exit(1)

    # ── Блок 4: пользователи (НЕкритичный — раньше именно тут всё падало) ──
    try:
        print("\n👤 Импортируем пользователей...")
        count_users = 0
        if "Accounts Balances" not in wb.sheetnames:
            print("   ⚠️  Лист 'Accounts Balances' не найден в файле — пропускаем (не критично).")
        else:
            for row in wb["Accounts Balances"].iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                username = str(row[0]).strip()
                role_str = str(row[1] or "user").strip().lower()
                balance = float(row[2] or 0)
                if not db.query(User).filter(User.username == username).first():
                    db.add(User(
                        username=username, full_name=username,
                        hashed_password=get_password_hash("changeme123"),
                        role=UserRole.ADMIN if role_str == "admin" else UserRole.OPERATOR,
                        balance=balance,
                    ))
                    count_users += 1
        db.flush()
        db.commit()  # ← ЧЕКПОИНТ 4
        print(f"   ✅ {count_users} пользователей (пароль по умолчанию: changeme123)")

    except Exception:
        db.rollback()
        print("\n⚠️  Не удалось импортировать пользователей из 'Accounts Balances' "
              "(это НЕ критично — склады/артикулы/остатки/движения уже сохранены):")
        traceback.print_exc()

    # ── Блок 5: admin-пользователь (отдельная мини-транзакция) ─────────────
    try:
        if not db.query(User).filter(User.username == settings.FIRST_ADMIN_USERNAME).first():
            db.add(User(
                username=settings.FIRST_ADMIN_USERNAME,
                full_name="Администратор",
                hashed_password=get_password_hash(settings.FIRST_ADMIN_PASSWORD),
                role=UserRole.ADMIN,
            ))
            db.commit()
            print(f"\n🔑 Создан admin: {settings.FIRST_ADMIN_USERNAME} / {settings.FIRST_ADMIN_PASSWORD}")
        else:
            print(f"\n🔑 Пользователь {settings.FIRST_ADMIN_USERNAME} уже существует — пропускаем.")
    except Exception:
        db.rollback()
        print("\n⚠️  Не удалось создать admin-пользователя (не критично — авторизация в проекте отключена):")
        traceback.print_exc()

    db.close()
    print("\n✅ Импорт завершён! (все успешно пройденные блоки сохранены в БД)")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", default="warehouse_report.xlsx")
    args = parser.parse_args()

    print("🔧 Инициализация БД...")
    init_db()
    print("✅ Таблицы созданы")
    seed(args.excel)
